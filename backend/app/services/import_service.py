from datetime import date, datetime, time
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.repositories import position_repository, schedule_repository
from app.schemas.imports import ImportRowErrorRead, RequirementsImportResultRead


def import_requirements_xlsx(db: Session, content: bytes) -> RequirementsImportResultRead:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid XLSX file format.") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid XLSX file format.")

    headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    required_headers = ["date", "position_id", "start_time", "end_time", "min_staff"]
    if any(header not in headers for header in required_headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid XLSX file format.")

    header_map = {header: headers.index(header) for header in required_headers}
    created_count = 0
    errors: list[ImportRowErrorRead] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None for cell in row):
            continue
        try:
            shift_date = _parse_date(row[header_map["date"]])
            position_id = _parse_int(row[header_map["position_id"]], "position_id")
            start_time = _parse_time(row[header_map["start_time"]], "start_time")
            end_time = _parse_time(row[header_map["end_time"]], "end_time")
            min_staff = _parse_int(row[header_map["min_staff"]], "min_staff")

            if end_time <= start_time:
                raise ValueError("end_time must be later than start_time.")
            if min_staff < 1:
                raise ValueError("min_staff must be greater than or equal to 1.")

            position = position_repository.get_position_by_id(db, position_id)
            if position is None:
                raise ValueError(f"Position {position_id} was not found.")

            schedule_repository.create_requirement(
                db,
                company_id=position.company_id,
                position_id=position_id,
                shift_date=shift_date,
                start_time=start_time,
                end_time=end_time,
                required_employees=min_staff,
            )
            created_count += 1
        except ValueError as exc:
            errors.append(ImportRowErrorRead(row=row_number, message=str(exc)))

    return RequirementsImportResultRead(created_count=created_count, errors=errors)


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError as exc:
            raise ValueError("date must be in YYYY-MM-DD format.") from exc
    raise ValueError("date is required.")


def _parse_time(value, field_name: str) -> time:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, str):
        try:
            return time.fromisoformat(value.strip())
        except ValueError as exc:
            raise ValueError(f"{field_name} must be in HH:MM[:SS] format.") from exc
    raise ValueError(f"{field_name} is required.")


def _parse_int(value, field_name: str) -> int:
    if value is None:
        raise ValueError(f"{field_name} is required.")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be an integer.") from exc
